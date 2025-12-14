import secrets
import string
import re
import io

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.tokens import default_token_generator
from django.contrib.admin.views.decorators import staff_member_required
from django.core.mail import send_mail
from django.core.serializers.json import DjangoJSONEncoder
from django.db import connection, models
from django.db.models import Q, Count, Sum, Avg, F
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.db.models.functions import TruncMonth
from django.conf import settings
from django.apps import apps

from django.contrib import messages

from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

from .models import *
from .forms import *
from .permissions import *
from .decorators import *


def help_page(request):
    """Страница справки"""
    return render(request, 'core/help.html')

def user_manual(request):
    """Руководство пользователя"""
    return render(request, 'core/user_manual.html')

def about_app(request):
    """О программе"""
    return render(request, 'core/about.html')

@login_required
def home(request):
    """Домашняя страница"""
    return render(request, 'core/home.html', {'title': 'Главная страница'})

@login_required
def dashboard(request):
    """Панель управления"""
    context = {
        'title': 'Панель управления',
        'user': request.user,
        'dish_count': Dish.objects.count(),
        'employee_count': Employee.objects.count(),
        'product_count': Product.objects.count(),
        'delivery_count': Delivery.objects.count(),
        'request_count': Request.objects.count(),
        'provider_count': Provider.objects.count(),
        'report_count': Report.objects.count(),
    }
    
    return render(request, 'core/dashboard.html', context)

def password_reset_request(request):
    """Смена пароля"""
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            try:
                user = User.objects.get(email=email)
                
                # Генерируем новый пароль
                new_password = ''.join(secrets.choice(string.ascii_letters + string.digits + '!@#$%^&*') for _ in range(12))
                
                # Устанавливаем новый пароль
                user.set_password(new_password)
                user.save()
                
                # Отправляем email с новым паролем
                subject = 'Восстановление пароля'
                message = f'''
                Здравствуйте,
                
                Ваш новый пароль: {new_password}
                
                Пожалуйста, измените его после входа в систему.
                
                С уважением,
                Администрация сайта
                '''
                
                from_email = settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@localhost'
                
                send_mail(
                    subject,
                    message,
                    from_email,
                    [email],
                    fail_silently=False,
                )
                
                messages.success(request, 'Новый пароль отправлен на ваш email. Пожалуйста, проверьте почту.')
                return redirect('login')
                
            except User.DoesNotExist:
                messages.error(request, 'Пользователь с таким email не найден')
    else:
        form = PasswordResetForm()
    
    return render(request, 'core/password_reset.html', {'form': form})

class TableListView(LoginRequiredMixin, ListView):
    """Базовый класс для отображения таблиц"""
    template_name = 'core/table_list.html'
    paginate_by = 10
    model = None
    table_title = ''
    fields_to_display = []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['table_title'] = self.table_title
        context['fields_to_display'] = self.fields_to_display
        context['model_name'] = self.model._meta.model_name if self.model else ''
        context['has_add_permission'] = self.request.user.has_perm(f'core.add_{self.model._meta.model_name}')
        context['has_change_permission'] = self.request.user.has_perm(f'core.change_{self.model._meta.model_name}')
        context['has_delete_permission'] = self.request.user.has_perm(f'core.delete_{self.model._meta.model_name}')
        return context
    
    def get_queryset(self):
        return self.model.objects.all()

@login_required
def director_tables(request):
    """Список всех таблиц для директора"""
    if not request.user.groups.filter(name='Директор').exists() and not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к этой странице')
        return redirect('dashboard')
    
    
    app_models = apps.get_app_config('core').get_models()
    
    tables = []
    for model in app_models:
        model_name = model._meta.model_name
        verbose_name = model._meta.verbose_name_plural
        
        if model_name in ['abbreviationtype', 'gendertype', 'eventtype']:
            continue
            
        try:
            count = model.objects.count()
        except:
            count = 0
        
        url_name = f'table_{model_name}'
        
        tables.append({
            'name': model_name,
            'verbose_name': verbose_name,
            'count': count,
            'url': url_name,
            'has_view': request.user.has_perm(f'core.view_{model_name}'),
        })
    
    return render(request, 'core/role_tables.html', {
        'title': 'Все таблицы',
        'role': 'Директор',
        'tables': sorted(tables, key=lambda x: x['verbose_name']),
        'description': 'Полный доступ ко всем таблицам системы'
    })


@login_required
def manager_tables(request):
    """Список таблиц для менеджера"""
    if not request.user.groups.filter(name='Менеджер').exists() and not request.user.is_superuser:
        if not request.user.groups.filter(name='Директор').exists():
            messages.error(request, 'У вас нет доступа к этой странице')
            return redirect('dashboard')
    
    # Таблицы для менеджера
    manager_models = [
        ('dish', 'Блюда', Dish),
        ('ingredient', 'Ингредиенты', Ingredient),
        ('request', 'Заявки', Request),
        ('requestproduct', 'Продукты в заявках', RequestProduct),
        ('delivery', 'Поставки', Delivery),
        ('deliveryproduct', 'Продукты в поставках', DeliveryProduct),
        ('product', 'Продукты', Product),
        ('provider', 'Поставщики', Provider),
        ('report', 'Отчеты', Report),
        ('reportdish', 'Блюда в отчетах', ReportDish),
        ('bank', 'Банки', Bank),
        ('division', 'Подразделения', Division),
    ]
    
    tables = []
    for model_name, verbose_name, model in manager_models:
        try:
            count = model.objects.count()
            url_name = f'table_{model_name}'
            
            tables.append({
                'name': model_name,
                'verbose_name': verbose_name,
                'count': count,
                'url': url_name,
                'has_view': request.user.has_perm(f'core.view_{model_name}'),
            })
        except:
            continue
    
    reference_models = [
        ('country', 'Страны', Country),
        ('city', 'Города', City),
        ('street', 'Улицы', Street),
        ('unitofmeasurement', 'Единицы измерения', UnitOfMeasurement),
        ('assortmentgroup', 'Группы ассортимента', AssortmentGroup),
    ]
    
    for model_name, verbose_name, model in reference_models:
        try:
            count = model.objects.count()
            url_name = f'table_{model_name}'
            
            tables.append({
                'name': model_name,
                'verbose_name': verbose_name,
                'count': count,
                'url': url_name,
                'has_view': request.user.has_perm(f'core.view_{model_name}'),
            })
        except:
            continue
    
    return render(request, 'core/role_tables.html', {
        'title': 'Таблицы менеджера',
        'role': 'Менеджер',
        'tables': sorted(tables, key=lambda x: x['verbose_name']),
        'description': 'Управление меню, поставками и заявками'
    })


@login_required
def chef_tables(request):
    """Список таблиц для шеф-повара"""
    if not request.user.groups.filter(name='Шеф-повар').exists() and not request.user.is_superuser:
        if not request.user.groups.filter(name='Директор').exists():
            messages.error(request, 'У вас нет доступа к этой странице')
            return redirect('dashboard')
    
    chef_models = [
        ('dish', 'Блюда', Dish),
        ('product', 'Продукты', Product),
        ('ingredient', 'Ингредиенты', Ingredient),
    ]
    
    tables = []
    for model_name, verbose_name, model in chef_models:
        try:
            count = model.objects.count()
            url_name = f'table_{model_name}'
            
            tables.append({
                'name': model_name,
                'verbose_name': verbose_name,
                'count': count,
                'url': url_name,
                'has_view': request.user.has_perm(f'core.view_{model_name}'),
            })
        except:
            continue
    
    reference_models = [
        ('assortmentgroup', 'Группы ассортимента', AssortmentGroup),
        ('unitofmeasurement', 'Единицы измерения', UnitOfMeasurement),
        ('country', 'Страны', Country),
        ('city', 'Города', City),
        ('street', 'Улицы', Street),
    ]
    
    for model_name, verbose_name, model in reference_models:
        try:
            count = model.objects.count()
            url_name = f'table_{model_name}'
            
            tables.append({
                'name': model_name,
                'verbose_name': verbose_name,
                'count': count,
                'url': url_name,
                'has_view': request.user.has_perm(f'core.view_{model_name}'),
            })
        except:
            continue
    
    return render(request, 'core/role_tables.html', {
        'title': 'Таблицы шеф-повара',
        'role': 'Шеф-повар',
        'tables': sorted(tables, key=lambda x: x['verbose_name']),
        'description': 'Просмотр меню и продуктов, редактирование ингредиентов'
    })


@login_required
def hr_tables(request):
    """Список таблиц для менеджера по кадрам"""
    if not request.user.groups.filter(name='Менеджер по кадрам').exists() and not request.user.is_superuser:
        # Проверяем, может пользователь директор?
        if not request.user.groups.filter(name='Директор').exists():
            messages.error(request, 'У вас нет доступа к этой странице')
            return redirect('dashboard')

    hr_models = [
        ('employee', 'Сотрудники', Employee),
        ('position', 'Должности', Position),
        ('placeofwork', 'Места работы', PlaceOfWork),
        ('department', 'Подразделения', Department),
        ('profession', 'Профессии', Profession),
        ('specialization', 'Специализации', Specialization),
        ('classification', 'Классификации', Classification),
        ('workbook', 'Трудовые книжки', WorkBook),
    ]
    
    tables = []
    for model_name, verbose_name, model in hr_models:
        try:
            count = model.objects.count()
            url_name = f'table_{model_name}'
            
            tables.append({
                'name': model_name,
                'verbose_name': verbose_name,
                'count': count,
                'url': url_name,
                'has_view': request.user.has_perm(f'core.view_{model_name}'),
            })
        except:
            continue
    
    reference_models = [
        ('country', 'Страны', Country),
        ('city', 'Города', City),
        ('street', 'Улицы', Street),
    ]
    
    for model_name, verbose_name, model in reference_models:
        try:
            count = model.objects.count()
            url_name = f'table_{model_name}'
            
            tables.append({
                'name': model_name,
                'verbose_name': verbose_name,
                'count': count,
                'url': url_name,
                'has_view': request.user.has_perm(f'core.view_{model_name}'),
            })
        except:
            continue
    
    return render(request, 'core/role_tables.html', {
        'title': 'Таблицы менеджера по кадрам',
        'role': 'Менеджер по кадрам',
        'tables': sorted(tables, key=lambda x: x['verbose_name']),
        'description': 'Управление персоналом и кадровыми данными'
    })


class UniversalTableView(LoginRequiredMixin, ListView):
    """Универсальный View для отображения всех таблиц"""
    template_name = 'core/universal_table.html'
    paginate_by = 10  # Базовое значение
    model = None
    
    def get_paginate_by(self, queryset):
        """Возвращает количество элементов на странице в зависимости от параметра per_page"""
        per_page = self.request.GET.get('per_page', 10)
        try:
            per_page = int(per_page)
            # Ограничиваем возможные значения
            if per_page not in [10, 20, 50, 100]:
                per_page = 10
        except ValueError:
            per_page = 10
        return per_page
    
    def get_queryset(self):

        queryset = super().get_queryset()
        
        # Получаем поисковый запрос
        search_query = self.request.GET.get('search', '').strip()
        
        if search_query and self.model:
            # Если есть поисковый запрос, фильтруем вручную
            filtered_objects = []
            
            # Получаем все объекты модели
            all_objects = self.model.objects.all()
            
            for obj in all_objects:
                # Собираем строку из всех значений полей объекта
                record_string = ""
                
                for field in self.model._meta.get_fields():
                    if not hasattr(field, 'get_internal_type'):
                        continue
                    
                    field_name = field.name
                    
                    # Пропускаем ManyToMany поля и обратные связи
                    if field.many_to_many or field.auto_created:
                        continue
                    
                    try:
                        # Получаем значение поля
                        field_value = getattr(obj, field_name)
                        
                        if field_value is not None:
                            # Для дат и времени конвертируем в строку в формате DD.MM.YYYY
                            if hasattr(field_value, 'strftime'):  # Это дата/время
                                if hasattr(field_value, 'date'):
                                    field_value = field_value.strftime('%d.%m.%Y')
                                else:
                                    field_value = str(field_value)
                            
                            # Для ForeignKey получаем строковое представление связанного объекта
                            elif hasattr(field_value, '_meta'):  # Это связанная модель
                                field_value = str(field_value)
                            
                            # Добавляем значение к строке
                            record_string += " " + str(field_value)
                    except:
                        # Если не удается получить значение, пропускаем поле
                        continue
                
                # Проверяем, содержится ли поисковый запрос в строке записи
                if search_query.lower() in record_string.lower():
                    filtered_objects.append(obj)
            
            # Возвращаем отфильтрованный QuerySet
            if filtered_objects:
                object_ids = [obj.id for obj in filtered_objects]
                queryset = self.model.objects.filter(id__in=object_ids)
            else:
                queryset = self.model.objects.none()
        
        # Применяем сортировку
        order_by = self.request.GET.get('order_by', '')
        direction = self.request.GET.get('direction', 'asc')  # 'asc' или 'desc'
        
        if order_by:
            # Проверяем, существует ли поле
            try:
                field_exists = False
                for field in self.model._meta.get_fields():
                    if field.name == order_by and not field.many_to_many and not field.auto_created:
                        field_exists = True
                        break
                
                if field_exists:
                    if direction == 'desc':
                        order_by_field = f'-{order_by}'
                    else:
                        order_by_field = order_by
                    
                    queryset = queryset.order_by(order_by_field)
            except:
                # Если поле не существует или ошибка сортировки, не сортируем
                pass
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.model:
            # Получаем название таблицы
            context['table_title'] = self.model._meta.verbose_name_plural
            
            # Получаем все поля модели, кроме id и полей ManyToMany
            fields = []
            for field in self.model._meta.get_fields():
                # Исключаем:
                # 1. Поля ManyToMany (они отображаются через related_name)
                # 2. Поле id (автоинкремент)
                # 3. Обратные связи
                if (not field.auto_created and 
                    field.name != 'id' and 
                    not field.many_to_many and
                    not isinstance(field, models.ManyToManyField)):
                    
                    # Для ForeignKey получаем связанный объект
                    if isinstance(field, models.ForeignKey):
                        fields.append({
                            'name': field.name,
                            'verbose_name': field.verbose_name,
                            'type': 'foreign_key',
                            'related_model': field.related_model,
                        })
                    # Для DateTimeField
                    elif isinstance(field, models.DateTimeField):
                        fields.append({
                            'name': field.name,
                            'verbose_name': field.verbose_name,
                            'type': 'datetime',
                        })
                    # Для DateField
                    elif isinstance(field, models.DateField):
                        fields.append({
                            'name': field.name,
                            'verbose_name': field.verbose_name,
                            'type': 'date',
                        })
                    # Для BooleanField
                    elif isinstance(field, models.BooleanField):
                        fields.append({
                            'name': field.name,
                            'verbose_name': field.verbose_name,
                            'type': 'boolean',
                        })
                    # Для ImageField/FileField
                    elif isinstance(field, models.ImageField) or isinstance(field, models.FileField):
                        fields.append({
                            'name': field.name,
                            'verbose_name': field.verbose_name,
                            'type': 'image',
                        })
                    # Для DecimalField
                    elif isinstance(field, models.DecimalField):
                        fields.append({
                            'name': field.name,
                            'verbose_name': field.verbose_name,
                            'type': 'decimal',
                            'max_digits': field.max_digits,
                            'decimal_places': field.decimal_places,
                        })
                    # Для ChoiceField
                    elif field.choices:
                        fields.append({
                            'name': field.name,
                            'verbose_name': field.verbose_name,
                            'type': 'choice',
                        })
                    # Для остальных полей (CharField, TextField и т.д.)
                    else:
                        fields.append({
                            'name': field.name,
                            'verbose_name': field.verbose_name,
                            'type': 'text',
                        })
            
            context['fields'] = fields
            context['model_name'] = self.model._meta.model_name
            context['search_query'] = self.request.GET.get('search', '')
            context['current_order_by'] = self.request.GET.get('order_by', '')
            context['current_direction'] = self.request.GET.get('direction', 'asc')
            
            # Проверяем права доступа
            context['has_add_permission'] = self.request.user.has_perm(f'core.add_{self.model._meta.model_name}')
            context['has_change_permission'] = self.request.user.has_perm(f'core.change_{self.model._meta.model_name}')
            context['has_delete_permission'] = self.request.user.has_perm(f'core.delete_{self.model._meta.model_name}')
        
        return context
    

class UniversalCreateView(LoginRequiredMixin, CreateView):
    """Универсальный View для создания записей"""
    template_name = 'core/create_form.html'
    model = None
    fields = '__all__'
    form_class = None
    
    def get_form_class(self):
        # Для специфических моделей используем свои формы
        if self.model == Dish:
            return DishForm
        elif self.model == Report:
            return ReportForm
        elif self.model == Request:
            return RequestForm
        elif self.model == Delivery:
            return DeliveryForm
        elif self.model == WorkBook:
            return WorkBookForm
        elif self.model == Employee:
            return EmployeeForm
        # Для всех остальных моделей используем универсальный Form
        else:
            class DynamicForm(UniversalForm):
                class Meta:
                    model = self.model
                    fields = '__all__'
            return DynamicForm
    
    def get_success_url(self):
        # Проверяем, была ли нажата кнопка "сохранить и добавить еще"
        if 'action' in self.request.POST and self.request.POST['action'] == 'save_and_add':
            return self.request.path  # Остаемся на той же странице
        else:
            return reverse_lazy(f'table_{self.model._meta.model_name}')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Добавить {self.model._meta.verbose_name}'
        context['model_name'] = self.model._meta.model_name
        context['verbose_name'] = self.model._meta.verbose_name
        context['is_create'] = True
        return context
    
    def form_valid(self, form):
        # Сохраняем объект
        response = super().form_valid(form)
        
        # Обрабатываем ManyToMany связи
        if self.model == Dish:
            # Получаем выбранные ингредиенты из формы
            selected_ingredients = form.cleaned_data.get('ingredients', [])
            # Присваиваем их ManyToMany полю
            self.object.ingredients.set(selected_ingredients)
        
        # Проверяем, была ли нажата кнопка "сохранить и добавить еще"
        if 'action' in self.request.POST and self.request.POST['action'] == 'save_and_add':
            # Если да, то перенаправляем на ту же страницу создания
            return redirect(self.request.path)
        
        return response
    
    def dispatch(self, request, *args, **kwargs):
        # Проверяем права на добавление
        if not request.user.has_perm(f'core.add_{self.model._meta.model_name}'):
            messages.error(request, 'У вас нет прав для добавления записей в эту таблицу')
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

class UniversalUpdateView(LoginRequiredMixin, UpdateView):
    """Универсальный View для редактирования записей"""
    template_name = 'core/create_form.html'
    model = None
    fields = '__all__'
    form_class = None
    
    def get_form_class(self):
        # Для специфических моделей используем свои формы
        if self.model == Dish:
            return DishForm
        elif self.model == Report:
            return ReportForm
        elif self.model == Request:
            return RequestForm
        elif self.model == Delivery:
            return DeliveryForm
        elif self.model == WorkBook:
            return WorkBookForm
        elif self.model == Employee:
            return EmployeeForm
        # Для всех остальных моделей используем универсальный Form
        else:
            class DynamicForm(UniversalForm):
                class Meta:
                    model = self.model
                    fields = '__all__'
            return DynamicForm
    
    def get_success_url(self):
        return reverse_lazy(f'table_{self.model._meta.model_name}')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Изменить {self.model._meta.verbose_name}'
        context['model_name'] = self.model._meta.model_name
        context['verbose_name'] = self.model._meta.verbose_name
        context['is_create'] = False
        return context
    
    def form_valid(self, form):
        # Сохраняем объект
        response = super().form_valid(form)
        
        # Обрабатываем ManyToMany связи
        if self.model == Dish:
            # Получаем выбранные ингредиенты из формы
            selected_ingredients = form.cleaned_data.get('ingredients', [])
            # Присваиваем их ManyToMany полю
            self.object.ingredients.set(selected_ingredients)
        
        # Для других моделей с ManyToMany полями добавьте обработку при необходимости
        
        return response
    
    def dispatch(self, request, *args, **kwargs):
        # Проверяем права на изменение
        if not request.user.has_perm(f'core.change_{self.model._meta.model_name}'):
            messages.error(request, 'У вас нет прав для изменения записей в этой таблице')
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Убедимся, что instance передается в форму
        if hasattr(self, 'object') and self.object:
            kwargs['instance'] = self.object
        return kwargs


class UniversalDeleteView(LoginRequiredMixin, DeleteView):
    """Универсальный View для удаления записей"""
    template_name = 'core/delete_confirm.html'
    model = None
    
    def get_success_url(self):
        return reverse_lazy(f'table_{self.model._meta.model_name}')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_name'] = self.model._meta.model_name
        context['verbose_name'] = self.model._meta.verbose_name
        return context
    
    def dispatch(self, request, *args, **kwargs):
        # Проверяем права на удаление
        if not request.user.has_perm(f'core.delete_{self.model._meta.model_name}'):
            messages.error(request, 'У вас нет прав для удаления записей из этой таблицы')
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)


@login_required
def sql_query_page(request):
    """Страница для выполнения SQL SELECT запросов"""
    user = request.user
    
    # Определяем доступные модели в зависимости от роли
    available_models = get_available_models_for_user(user)
    
    # Подготовим шаблоны запросов в зависимости от роли
    template_queries = get_template_queries_for_user(user)
    
    if request.method == 'POST':
        if 'export' in request.POST:
            # Обработка экспорта
            sql_query = request.POST.get('sql_query', '').strip()
            
            if not sql_query:
                messages.error(request, 'Введите SQL запрос для экспорта!')
                return render(request, 'core/sql_query.html', {
                    'available_models': available_models,
                    'template_queries': template_queries,
                    'sql_query': sql_query
                })
            
            return export_sql_results_to_excel(sql_query, user)
        
        # Обработка выполнения запроса
        sql_query = request.POST.get('sql_query', '').strip()
        
        # Проверяем, что запрос является SELECT
        if not is_valid_select_query(sql_query):
            messages.error(request, 'Разрешены только SELECT запросы!')
            return render(request, 'core/sql_query.html', {
                'available_models': available_models,
                'template_queries': template_queries,
                'sql_query': sql_query
            })
        
        # Проверяем, что запрос использует только разрешенные таблицы
        if not is_query_using_allowed_tables(sql_query, available_models):
            messages.error(request, 'Запрос использует запрещенные таблицы!')
            return render(request, 'core/sql_query.html', {
                'available_models': available_models,
                'template_queries': template_queries,
                'sql_query': sql_query
            })
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql_query)
                columns = [col[0] for col in cursor.description]
                results = cursor.fetchall()
                
                return render(request, 'core/sql_query.html', {
                    'available_models': available_models,
                    'template_queries': template_queries,
                    'sql_query': sql_query,
                    'columns': columns,
                    'results': results,
                    'query_executed': True
                })
        except Exception as e:
            messages.error(request, f'Ошибка выполнения запроса: {str(e)}')
            return render(request, 'core/sql_query.html', {
                'available_models': available_models,
                'template_queries': template_queries,
                'sql_query': sql_query
            })
    
    return render(request, 'core/sql_query.html', {
        'available_models': available_models,
        'template_queries': template_queries
    })


def export_sql_results_to_excel(sql_query, user):
    """Экспортирует результаты SQL запроса в Excel файл"""
    # Проверяем, что запрос является SELECT
    if not is_valid_select_query(sql_query):
        messages.error(request, 'Разрешены только SELECT запросы!')
        return HttpResponse('Разрешены только SELECT запросы!', status=400)
    
    # Определяем доступные модели для пользователя
    available_models = get_available_models_for_user(user)
    
    # Проверяем, что запрос использует только разрешенные таблицы
    if not is_query_using_allowed_tables(sql_query, available_models):
        messages.error(request, 'Запрос использует запрещенные таблицы!')
        return HttpResponse('Запрос использует запрещенные таблицы!', status=400)
    
    try:
        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            columns = [col[0] for col in cursor.description]
            results = cursor.fetchall()
            
            # Создаем Excel файл
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'SQL Results'
            
            # Добавляем заголовки
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_num, value=column_title)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Добавляем данные
            for row_num, row_data in enumerate(results, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_num, column=col_num, value=str(cell_value))
            
            # Автонастройка ширины колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Максимальная ширина 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Подготовка ответа
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=sql_results_{user.username}_{user.id}.xlsx'
            
            workbook.save(response)
            return response
            
    except Exception as e:
        return HttpResponse(f'Ошибка экспорта: {str(e)}', status=500)


def get_available_models_for_user(user):
    """Возвращает список доступных моделей в зависимости от роли пользователя"""
    models_list = []
    
    if user.groups.filter(name='Директор').exists() or user.is_superuser:
        # Директор имеет доступ ко всем моделям
        app_models = apps.get_app_config('core').get_models()
        for model in app_models:
            model_name = model._meta.model_name
            if model_name not in ['abbreviationtype', 'gendertype', 'eventtype']:
                if user.has_perm(f'core.view_{model_name}'):
                    models_list.append({
                        'name': model_name,
                        'verbose_name': model._meta.verbose_name_plural
                    })
    elif user.groups.filter(name='Менеджер').exists():
        # Менеджер имеет доступ к определенным моделям
        manager_models = [
            ('dish', 'Блюда', Dish),
            ('ingredient', 'Ингредиенты', Ingredient),
            ('request', 'Заявки', Request),
            ('requestproduct', 'Продукты в заявках', RequestProduct),
            ('delivery', 'Поставки', Delivery),
            ('deliveryproduct', 'Продукты в поставках', DeliveryProduct),
            ('product', 'Продукты', Product),
            ('provider', 'Поставщики', Provider),
            ('report', 'Отчеты', Report),
            ('reportdish', 'Блюда в отчетах', ReportDish),
            ('bank', 'Банки', Bank),
            ('division', 'Подразделения', Division),
            ('country', 'Страны', Country),
            ('city', 'Города', City),
            ('street', 'Улицы', Street),
            ('unitofmeasurement', 'Единицы измерения', UnitOfMeasurement),
            ('assortmentgroup', 'Группы ассортимента', AssortmentGroup),
        ]
        
        for model_name, verbose_name, model in manager_models:
            if user.has_perm(f'core.view_{model_name}'):
                models_list.append({
                    'name': model_name,
                    'verbose_name': verbose_name
                })
    elif user.groups.filter(name='Шеф-повар').exists():
        # Шеф-повар имеет доступ к определенным моделям
        chef_models = [
            ('dish', 'Блюда', Dish),
            ('product', 'Продукты', Product),
            ('ingredient', 'Ингредиенты', Ingredient),
            ('assortmentgroup', 'Группы ассортимента', AssortmentGroup),
            ('unitofmeasurement', 'Единицы измерения', UnitOfMeasurement),
            ('country', 'Страны', Country),
            ('city', 'Города', City),
            ('street', 'Улицы', Street),
        ]
        
        for model_name, verbose_name, model in chef_models:
            if user.has_perm(f'core.view_{model_name}'):
                models_list.append({
                    'name': model_name,
                    'verbose_name': verbose_name
                })
    elif user.groups.filter(name='Менеджер по кадрам').exists():
        # Менеджер по кадрам имеет доступ к определенным моделям
        hr_models = [
            ('employee', 'Сотрудники', Employee),
            ('position', 'Должности', Position),
            ('placeofwork', 'Места работы', PlaceOfWork),
            ('department', 'Подразделения', Department),
            ('profession', 'Профессии', Profession),
            ('specialization', 'Специализации', Specialization),
            ('classification', 'Классификации', Classification),
            ('workbook', 'Трудовые книжки', WorkBook),
            ('country', 'Страны', Country),
            ('city', 'Города', City),
            ('street', 'Улицы', Street),
        ]
        
        for model_name, verbose_name, model in hr_models:
            if user.has_perm(f'core.view_{model_name}'):
                models_list.append({
                    'name': model_name,
                    'verbose_name': verbose_name
                })
    
    return models_list


def get_template_queries_for_user(user):
    """Возвращает шаблоны SQL запросов в зависимости от роли пользователя"""
    templates = []
    
    if user.groups.filter(name='Шеф-повар').exists():
        templates.extend([
            {
                'name': 'Шаблон для вывода блюд определенной группы ассортимента',
                'query': 'SELECT * FROM core_dish WHERE assortment_group_id = (SELECT id FROM core_assortmentgroup WHERE name = \'Суп\')'
            },
            {
                'name': 'Шаблон для вывода блюд с определенным ингредиентом',
                'query': 'SELECT d.name, d.price, d.output FROM core_dish d JOIN core_dish_ingredients di ON d.id = di.dish_id JOIN core_ingredient i ON di.ingredient_id = i.id WHERE i.name LIKE \'%курица%\''
            },
            {
                'name': 'Шаблон для вывода всех блюд с ценой в определенном диапазоне',
                'query': 'SELECT * FROM core_dish WHERE price BETWEEN 100 AND 500 ORDER BY price'
            }
        ])
    elif user.groups.filter(name='Менеджер').exists():
        templates.extend([
            {
                'name': 'Шаблон для вывода всех поставок за последнюю неделю',
                'query': 'SELECT * FROM core_delivery WHERE date >= date(\'now\', \'-7 days\')'
            },
            {
                'name': 'Шаблон для вывода продуктов с низким остатком',
                'query': 'SELECT * FROM core_product WHERE remaining_stock < 10 ORDER BY remaining_stock ASC'
            },
            {
                'name': 'Шаблон для вывода всех заявок за последний месяц',
                'query': 'SELECT * FROM core_request WHERE date >= date(\'now\', \'-30 days\')'
            }
        ])
    elif user.groups.filter(name='Менеджер по кадрам').exists():
        templates.extend([
            {
                'name': 'Шаблон для вывода всех сотрудников определенной должности',
                'query': 'SELECT * FROM core_employee WHERE position_id = (SELECT id FROM core_position WHERE name = \'Повар\')'
            },
            {
                'name': 'Шаблон для вывода сотрудников по городу',
                'query': 'SELECT e.first_name, e.last_name, c.name as city FROM core_employee e JOIN core_city c ON e.city_id = c.id WHERE c.name = \'Москва\''
            },
            {
                'name': 'Шаблон для вывода сотрудников с определенной профессией',
                'query': 'SELECT * FROM core_employee e JOIN core_workbook w ON e.id = w.employee_id JOIN core_profession p ON w.profession_id = p.id WHERE p.name LIKE \'%повар%\''
            }
        ])
    elif user.groups.filter(name='Директор').exists() or user.is_superuser:
        templates.extend([
            {
                'name': 'Шаблон для вывода всех блюд с ценой выше средней',
                'query': 'SELECT * FROM core_dish WHERE price > (SELECT AVG(price) FROM core_dish)'
            },
            {
                'name': 'Шаблон для вывода топ-5 самых дорогих блюд',
                'query': 'SELECT * FROM core_dish ORDER BY price DESC LIMIT 5'
            },
            {
                'name': 'Шаблон для вывода сотрудников по должности',
                'query': 'SELECT e.first_name, e.last_name, p.name as position FROM core_employee e JOIN core_position p ON e.position_id = p.id WHERE p.name LIKE \'%повар%\''
            }
        ])
    
    return templates

def is_valid_select_query(sql_query):
    """Проверяет, что запрос является SELECT запросом"""
    sql_query = sql_query.strip().upper()
    return sql_query.startswith('SELECT ')


def is_query_using_allowed_tables(sql_query, allowed_models):
    """Проверяет, что запрос использует только разрешенные таблицы"""
    
    # Получаем список разрешенных имен таблиц (моделей)
    allowed_table_names = [model['name'] for model in allowed_models]
    
    # Ищем имена таблиц после FROM, JOIN и т.д.
    pattern = r'(?:FROM|JOIN|INNER\s+JOIN|LEFT\s+JOIN|RIGHT\s+JOIN|CROSS\s+JOIN)\s+([`"]?)([a-zA-Z_][a-zA-Z0-9_]*)(\1)(?:\s|$|[^a-zA-Z0-9_])'
    
    matches = re.findall(pattern, sql_query.upper())
    
    # Извлекаем только имена таблиц (второй элемент в кортеже)
    found_tables = []
    for match in matches:
        table_name = match[1]  # Второй элемент - это имя таблицы
        # Убираем префикс 'core_'
        clean_table = table_name.replace('CORE_', '')
        if clean_table:
            found_tables.append(clean_table.lower())  # Приводим к нижнему регистру
    
    # Проверяем каждую найденную таблицу
    for table in found_tables:
        if table not in allowed_table_names:
            return False
    
    return True


def miscellaneous_page(request):
    """Страница "Разное" """
    return render(request, 'core/miscellaneous.html')


def settings_page(request):
    """Страница настроек"""
    if request.method == 'POST':
        # Обработка настроек
        font_size = request.POST.get('font_size', 'normal')
        language = request.POST.get('language', 'ru')
        theme = request.POST.get('theme', 'light')  # Добавляем обработку темы
        
        # Сохраняем настройки в сессии
        request.session['font_size'] = font_size
        request.session['language'] = language
        request.session['theme'] = theme  # Сохраняем тему
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Если AJAX запрос, возвращаем JSON
            return JsonResponse({'status': 'success', 'message': 'Настройки сохранены!'})
        else:
            # Если обычный запрос, возвращаем редирект
            messages.success(request, 'Настройки сохранены!')
            return redirect('settings')
    
    # Загружаем текущие настройки
    font_size = request.session.get('font_size', 'normal')
    language = request.session.get('language', 'ru')
    theme = request.session.get('theme', 'light')  # Загружаем текущую тему
    
    context = {
        'font_size': font_size,
        'language': language,
        'theme': theme  # Передаем тему в контекст
    }
    return render(request, 'core/settings.html', context)





def change_password(request):
    """Страница смены пароля"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Важно для сохранения сессии
            messages.success(request, 'Ваш пароль успешно изменен!')
            return redirect('change_password')
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки.')
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'core/change_password.html', {
        'form': form
    })


@login_required
def analytics_dashboard(request):
    """Дашборд аналитики"""
    user = request.user
    
    # Получаем даты из GET параметров (для фильтрации при переходе с других страниц)
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    start_date = None
    end_date = None
    
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    
    # Если даты не переданы, используем последние 30 дней
    if not start_date:
        start_date = (datetime.now() - timedelta(days=30)).date()
    if not end_date:
        end_date = datetime.now().date()
    
    # Определяем доступные данные в зависимости от роли
    available_data = get_available_analytics_data(user)
    
    # Генерируем данные для графиков с фильтрацией по дате
    charts_data = generate_charts_data(user, start_date, end_date)
    
    context = {
        'charts_data': charts_data,
        'available_data': available_data,
        'date_range': {
            'start': start_date.strftime('%Y-%m-%d'),
            'end': end_date.strftime('%Y-%m-%d')
        }
    }
    
    return render(request, 'core/analytics.html', context)


def get_available_analytics_data(user):
    """Возвращает доступные данные для аналитики в зависимости от роли пользователя"""
    data = {
        'has_dish_data': user.has_perm('core.view_dish'),
        'has_product_data': user.has_perm('core.view_product'),
        'has_employee_data': user.has_perm('core.view_employee'),
        'has_report_data': user.has_perm('core.view_report'),
        'has_delivery_data': user.has_perm('core.view_delivery'),
        'has_request_data': user.has_perm('core.view_request'),
    }
    
    # Добавляем статистику
    if data['has_dish_data']:
        data['dish_count'] = Dish.objects.count()
    if data['has_product_data']:
        data['product_count'] = Product.objects.count()
    if data['has_employee_data']:
        data['employee_count'] = Employee.objects.count()
    if data['has_report_data']:
        data['report_count'] = Report.objects.count()
    
    return data


def generate_charts_data(user, start_date=None, end_date=None):
    """Генерирует данные для графиков с возможностью фильтрации по дате"""
    charts_data = []
    
    # Создаем фильтр по дате, если они указаны
    date_filter = {}
    if start_date:
        date_filter['date__gte'] = start_date
    if end_date:
        date_filter['date__lte'] = end_date
    
    # График: Топ-5 блюд по количеству
    if user.has_perm('core.view_dish') and user.has_perm('core.view_reportdish'):
        try:
            # Получаем данные из отчетов о блюдах
            # Фильтруем по дате отчетов
            top_dishes_query = ReportDish.objects.select_related('report').values('dish__name')
            
            if start_date or end_date:
                top_dishes_query = top_dishes_query.filter(report__date__range=[start_date, end_date])
            
            top_dishes = top_dishes_query.annotate(
                total_quantity=Sum('quantity')
            ).order_by('-total_quantity')[:5]
            
            if top_dishes:
                dish_names = [item['dish__name'] for item in top_dishes if item['dish__name']]
                quantities = [float(item['total_quantity'] or 0) for item in top_dishes]
                
                charts_data.append({
                    'title': 'Топ-5 блюд по количеству продаж',
                    'chart_id': 'top_dishes_chart',
                    'chart_type': 'bar',
                    'x_data': dish_names,
                    'y_data': quantities,
                    'x_label': 'Блюда',
                    'y_label': 'Количество',
                    'color': 'skyblue'
                })
        except Exception as e:
            print(f"Error generating top dishes chart: {e}")
    
    # График: Выручка по группам ассортимента
    if user.has_perm('core.view_dish') and user.has_perm('core.view_reportdish'):
        try:
            # Получаем данные о выручке по группам ассортимента
            sales_by_group_query = ReportDish.objects.select_related('dish', 'dish__assortment_group', 'report').values(
                'dish__assortment_group__name'
            )
            
            if start_date or end_date:
                sales_by_group_query = sales_by_group_query.filter(report__date__range=[start_date, end_date])
            
            sales_by_group = sales_by_group_query.annotate(
                total_revenue=Sum(F('quantity') * F('dish__price'))
            ).order_by('-total_revenue')
            
            if sales_by_group:
                group_names = [item['dish__assortment_group__name'] for item in sales_by_group if item['dish__assortment_group__name']]
                revenues = [float(item['total_revenue'] or 0) for item in sales_by_group]
                
                charts_data.append({
                    'title': 'Выручка по группам ассортимента',
                    'chart_id': 'revenue_by_group_chart',
                    'chart_type': 'bar',
                    'x_data': group_names,
                    'y_data': revenues,
                    'x_label': 'Группы ассортимента',
                    'y_label': 'Выручка (₽)',
                    'color': 'lightgreen'
                })
        except Exception as e:
            print(f"Error generating revenue by group chart: {e}")
    
    # График: Остатки продуктов
    if user.has_perm('core.view_product'):
        try:
            low_stock_products = Product.objects.filter(
                remaining_stock__lt=20
            ).order_by('remaining_stock')[:10]
            
            if low_stock_products:
                product_names = [p.name for p in low_stock_products]
                stock_levels = [float(p.remaining_stock) for p in low_stock_products]
                
                charts_data.append({
                    'title': 'Продукты с низким остатком',
                    'chart_id': 'low_stock_chart',
                    'chart_type': 'bar',
                    'x_data': product_names,
                    'y_data': stock_levels,
                    'x_label': 'Продукты',
                    'y_label': 'Остаток',
                    'color': 'orange'
                })
        except Exception as e:
            print(f"Error generating low stock chart: {e}")
    
    # График: Поставки по месяцам
    if user.has_perm('core.view_delivery'):
        try:
            
            
            # Применяем фильтр по дате к поставкам
            deliveries_query = Delivery.objects.all()
            
            if start_date:
                deliveries_query = deliveries_query.filter(date__gte=start_date)
            if end_date:
                deliveries_query = deliveries_query.filter(date__lte=end_date)
            
            # Если не заданы даты, берем последние 12 месяцев
            if not start_date and not end_date:
                start_date_calc = datetime.now() - timedelta(days=365)
                deliveries_query = deliveries_query.filter(date__gte=start_date_calc)
            
            # Группируем поставки по месяцам
            monthly_deliveries = deliveries_query.annotate(
                month=TruncMonth('date')
            ).values('month').annotate(
                delivery_count=Count('id')
            ).order_by('month')
            
            if monthly_deliveries:
                months = [item['month'].strftime('%Y-%m') for item in monthly_deliveries]
                counts = [item['delivery_count'] for item in monthly_deliveries]
                
                charts_data.append({
                    'title': 'Количество поставок по месяцам',
                    'chart_id': 'monthly_deliveries_chart',
                    'chart_type': 'line',
                    'x_data': months,
                    'y_data': counts,
                    'x_label': 'Месяц',
                    'y_label': 'Количество поставок',
                    'color': 'purple'
                })
        except Exception as e:
            print(f"Error generating deliveries chart: {e}")
    
    # График: Средняя цена блюд по группам
    if user.has_perm('core.view_dish'):
        try:
            # Фильтрация по дате не применяется к блюдам, так как это статические данные
            avg_price_by_group = Dish.objects.values(
                'assortment_group__name'
            ).annotate(
                avg_price=Avg('price')
            ).order_by('-avg_price')
            
            if avg_price_by_group:
                group_names = [item['assortment_group__name'] for item in avg_price_by_group if item['assortment_group__name']]
                avg_prices = [float(item['avg_price'] or 0) for item in avg_price_by_group]
                
                charts_data.append({
                    'title': 'Средняя цена блюд по группам',
                    'chart_id': 'avg_price_chart',
                    'chart_type': 'bar',
                    'x_data': group_names,
                    'y_data': avg_prices,
                    'x_label': 'Группы ассортимента',
                    'y_label': 'Средняя цена (₽)',
                    'color': 'coral'
                })
        except Exception as e:
            print(f"Error generating avg price chart: {e}")
    
    return charts_data


@csrf_exempt
def update_charts_data(request):
    """Обновление данных графиков через AJAX"""
    if request.method == 'POST':
        try:
            start_date_str = request.POST.get('start_date')
            end_date_str = request.POST.get('end_date')
            
            start_date = None
            end_date = None
            
            if start_date_str:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            if end_date_str:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
                
            user = request.user
            charts_data = generate_charts_data(user, start_date, end_date)
            
            return JsonResponse({'charts_data': charts_data}, encoder=DjangoJSONEncoder)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


def export_chart(request, chart_id):
    """Экспорт графика в PNG"""
    
    
    fig, ax = plt.subplots(figsize=(10, 6))
    

    if chart_id == 'top_dishes_chart':
        try:
            top_dishes = ReportDish.objects.values('dish__name').annotate(
                total_quantity=Sum('quantity')
            ).order_by('-total_quantity')[:5]
            
            if top_dishes:
                dish_names = [item['dish__name'] for item in top_dishes if item['dish__name']]
                quantities = [float(item['total_quantity'] or 0) for item in top_dishes]
                
                ax.bar(dish_names, quantities, color='skyblue')
                ax.set_title('Топ-5 блюд по количеству продаж')
                ax.set_xlabel('Блюда')
                ax.set_ylabel('Количество')
                plt.xticks(rotation=45, ha='right')
            else:
                ax.text(0.5, 0.5, 'Нет данных', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
                ax.set_title('Нет данных для отображения')
        except Exception as e:
            ax.text(0.5, 0.5, 'Ошибка: ' + str(e), horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
            ax.set_title('Ошибка загрузки данных')
    
    elif chart_id == 'revenue_by_group_chart':
        try:
            sales_by_group = ReportDish.objects.select_related('dish', 'dish__assortment_group').values(
                'dish__assortment_group__name'
            ).annotate(
                total_revenue=Sum(F('quantity') * F('dish__price'))
            ).order_by('-total_revenue')
            
            if sales_by_group:
                group_names = [item['dish__assortment_group__name'] for item in sales_by_group if item['dish__assortment_group__name']]
                revenues = [float(item['total_revenue'] or 0) for item in sales_by_group]
                
                ax.bar(group_names, revenues, color='lightgreen')
                ax.set_title('Выручка по группам ассортимента')
                ax.set_xlabel('Группы ассортимента')
                ax.set_ylabel('Выручка (₽)')
                plt.xticks(rotation=45, ha='right')
            else:
                ax.text(0.5, 0.5, 'Нет данных', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
                ax.set_title('Нет данных для отображения')
        except Exception as e:
            ax.text(0.5, 0.5, 'Ошибка: ' + str(e), horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
            ax.set_title('Ошибка загрузки данных')
    
    elif chart_id == 'low_stock_chart':
        try:
            low_stock_products = Product.objects.filter(
                remaining_stock__lt=20
            ).order_by('remaining_stock')[:10]
            
            if low_stock_products:
                product_names = [p.name for p in low_stock_products]
                stock_levels = [float(p.remaining_stock) for p in low_stock_products]
                
                ax.bar(product_names, stock_levels, color='orange')
                ax.set_title('Продукты с низким остатком')
                ax.set_xlabel('Продукты')
                ax.set_ylabel('Остаток')
                plt.xticks(rotation=45, ha='right')
            else:
                ax.text(0.5, 0.5, 'Нет данных', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
                ax.set_title('Нет данных для отображения')
        except Exception as e:
            ax.text(0.5, 0.5, 'Ошибка: ' + str(e), horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
            ax.set_title('Ошибка загрузки данных')
    
    elif chart_id == 'monthly_deliveries_chart':
        try:
            
            
            start_date = datetime.now() - timedelta(days=365)
            
            # Группируем поставки по месяцам
            monthly_deliveries = Delivery.objects.filter(
                date__gte=start_date
            ).annotate(
                month=TruncMonth('date')
            ).values('month').annotate(
                delivery_count=Count('id')
            ).order_by('month')
            
            if monthly_deliveries:
                months = [item['month'].strftime('%Y-%m') for item in monthly_deliveries]
                counts = [item['delivery_count'] for item in monthly_deliveries]
                
                ax.plot(months, counts, marker='o', linewidth=2, color='purple')
                ax.set_title('Количество поставок по месяцам')
                ax.set_xlabel('Месяц')
                ax.set_ylabel('Количество поставок')
                plt.xticks(rotation=45, ha='right')
            else:
                ax.text(0.5, 0.5, 'Нет данных', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
                ax.set_title('Нет данных для отображения')
        except Exception as e:
            ax.text(0.5, 0.5, 'Ошибка: ' + str(e), horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
            ax.set_title('Ошибка загрузки данных')
    
    elif chart_id == 'avg_price_chart':
        try:
            avg_price_by_group = Dish.objects.values(
                'assortment_group__name'
            ).annotate(
                avg_price=Avg('price')
            ).order_by('-avg_price')
            
            if avg_price_by_group:
                group_names = [item['assortment_group__name'] for item in avg_price_by_group if item['assortment_group__name']]
                avg_prices = [float(item['avg_price'] or 0) for item in avg_price_by_group]
                
                ax.bar(group_names, avg_prices, color='coral')
                ax.set_title('Средняя цена блюд по группам ассортимента')
                ax.set_xlabel('Группы ассортимента')
                ax.set_ylabel('Средняя цена (₽)')
                plt.xticks(rotation=45, ha='right')
            else:
                ax.text(0.5, 0.5, 'Нет данных', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
                ax.set_title('Нет данных для отображения')
        except Exception as e:
            ax.text(0.5, 0.5, 'Ошибка: ' + str(e), horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
            ax.set_title('Ошибка загрузки данных')
    
    else:
        ax.text(0.5, 0.5, f'График: {chart_id}', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
        ax.set_title(f'График: {chart_id}')
    
    plt.tight_layout()
    
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='image/png')
    response['Content-Disposition'] = f'attachment; filename={chart_id}.png'
    
    plt.close()
    return response