from django.http import HttpResponse
from django.db import connection
from django.apps import apps
from django.contrib.auth import get_user_model

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import re

from core.models import (
    Dish, Ingredient, Request, RequestProduct, Delivery, DeliveryProduct,
    Product, Provider, Report, ReportDish, Bank, Division, Country,
    City, Street, UnitOfMeasurement, AssortmentGroup, Employee,
    Position, PlaceOfWork, Department, Profession, Specialization,
    Classification, WorkBook
)

User = get_user_model()

def export_sql_results_to_excel(sql_query, user):
    """Экспортирует результаты SQL запроса в Excel файл"""
    # Проверяем, что запрос является SELECT
    if not is_valid_select_query(sql_query):
        return HttpResponse('Разрешены только SELECT запросы!', status=400)
    
    available_models = get_available_models_for_user(user)
    
    if not is_query_using_allowed_tables(sql_query, available_models):
        return HttpResponse('Запрос использует запрещенные таблицы!', status=400)
    
    try:
        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            columns = [col[0] for col in cursor.description]
            results = cursor.fetchall()
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'SQL Results'
            
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_num, value=column_title)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            for row_num, row_data in enumerate(results, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_num, column=col_num, value=str(cell_value))
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=sql_results_{user.username}_{user.id}.xlsx'
            
            workbook.save(response)
            return response
            
    except Exception as e:
        return HttpResponse(f'Ошибка экспорта: {str(e)}', status=500)

def get_available_models_for_user(user):
    """Возвращает список доступных моделей в зависимости от роли пользователя"""
    models_list = []
    
    if user.groups.filter(name='Директор').exists() or user.is_superuser:
        # Для директора - все таблицы приложения core, исключая системные Django таблицы
        from django.apps import apps
        
        # Получаем все модели из приложения core
        core_app = apps.get_app_config('core')
        core_models = core_app.get_models()
        
        for model in core_models:
            model_name = model._meta.model_name
            
            # Исключаем системные таблицы Django и справочные таблицы
            excluded_names = [
                'abbreviationtype', 'gendertype', 'eventtype',  # Старые справочные таблицы
                'permission', 'group', 'user', 'contenttype', 'session', 'adminlogentry',
                'migration', 'logentry', 'tokenproxy', 'permissionproxy', 'groupproxy',
                'userproxy', 'contenttypeproxy', 'sessionproxy', 'adminlogentryproxy',
                'migrationproxy', 'logentryproxy'
            ]
            
            # Также исключаем таблицы, начинающиеся с auth_, sessions_, admin_, contenttypes_
            if (model_name not in excluded_names and 
                not model_name.startswith(('auth_', 'session_', 'admin_', 'contenttype_')) and
                not model_name.startswith(('logentry_', 'permission_', 'group_', 'user_'))):
                
                # Проверяем права на просмотр (если пользователь не суперпользователь)
                if user.is_superuser or user.has_perm(f'core.view_{model_name}'):
                    models_list.append({
                        'name': model_name,
                        'verbose_name': model._meta.verbose_name_plural
                    })
        many_to_many_intermediate_tables = [
            {'name': 'dish_ingredients', 'verbose_name': 'Связь блюдо-ингредиент'},
        ]
        
        for table in many_to_many_intermediate_tables:
            models_list.append({
                'name': table['name'],
                'verbose_name': table['verbose_name']
            })
    elif user.groups.filter(name='Менеджер').exists():
        # Остальная логика для менеджера остается без изменений
        manager_models = [
            ('dish', 'Блюда', Dish),
            ('ingredient', 'Ингредиенты', Ingredient),
            ('request', 'Заявки', Request),
            ('requestproduct', 'Продукты в заявках', RequestProduct),
            ('delivery', 'Поставки', Delivery),
            ('deliveryproduct', 'Продукты в поставках', DeliveryProduct),
            ('product', 'Продукты', Product),
            ('provider', 'Поставщики', Provider),
            ('report', 'Отчеты', Report),
            ('reportdish', 'Блюда в отчетах', ReportDish),
            ('bank', 'Банки', Bank),
            ('division', 'Подразделения', Division),
            ('country', 'Страны', Country),
            ('city', 'Города', City),
            ('street', 'Улицы', Street),
            ('unitofmeasurement', 'Единицы измерения', UnitOfMeasurement),
            ('assortmentgroup', 'Группы ассортимента', AssortmentGroup),
        ]
        
        for model_name, verbose_name, model in manager_models:
            if user.has_perm(f'core.view_{model_name}'):
                models_list.append({
                    'name': model_name,
                    'verbose_name': verbose_name
                })
    elif user.groups.filter(name='Шеф-повар').exists():
        # Остальная логика для шеф-повара остается без изменений
        chef_models = [
            ('dish', 'Блюда', Dish),
            ('product', 'Продукты', Product),
            ('ingredient', 'Ингредиенты', Ingredient),
            ('assortmentgroup', 'Группы ассортимента', AssortmentGroup),
            ('unitofmeasurement', 'Единицы измерения', UnitOfMeasurement),
            ('country', 'Страны', Country),
            ('city', 'Города', City),
            ('street', 'Улицы', Street),
        ]
        
        for model_name, verbose_name, model in chef_models:
            if user.has_perm(f'core.view_{model_name}'):
                models_list.append({
                    'name': model_name,
                    'verbose_name': verbose_name
                })
    elif user.groups.filter(name='Менеджер по кадрам').exists():
        # Остальная логика для менеджера по кадрам остается без изменений
        hr_models = [
            ('employee', 'Сотрудники', Employee),
            ('position', 'Должности', Position),
            ('placeofwork', 'Места работы', PlaceOfWork),
            ('department', 'Подразделения', Department),
            ('profession', 'Профессии', Profession),
            ('specialization', 'Специализации', Specialization),
            ('classification', 'Классификации', Classification),
            ('workbook', 'Трудовые книжки', WorkBook),
            ('country', 'Страны', Country),
            ('city', 'Города', City),
            ('street', 'Улицы', Street),
        ]
        
        for model_name, verbose_name, model in hr_models:
            if user.has_perm(f'core.view_{model_name}'):
                models_list.append({
                    'name': model_name,
                    'verbose_name': verbose_name
                })
    
    return models_list

def get_template_queries_for_user(user):
    """Возвращает шаблоны SQL запросов в зависимости от роли пользователя"""
    templates = []
    
    if user.groups.filter(name='Шеф-повар').exists():
        templates.extend([
            {
                'name': 'Шаблон для вывода блюд определенной группы ассортимента',
                'query': 'SELECT * FROM core_dish WHERE assortment_group_id = (SELECT id FROM core_assortmentgroup WHERE name = \'Суп\')'
            },
            {
                'name': 'Шаблон для вывода блюд с определенным ингредиентом',
                'query': 'SELECT d.name, d.price, d.output FROM core_dish d JOIN core_dish_ingredients di ON d.id = di.dish_id JOIN core_ingredient i ON di.ingredient_id = i.id WHERE i.name LIKE \'%курица%\''
            },
            {
                'name': 'Шаблон для вывода всех блюд с ценой в определенном диапазоне',
                'query': 'SELECT * FROM core_dish WHERE price BETWEEN 100 AND 500 ORDER BY price'
            }
        ])
    elif user.groups.filter(name='Менеджер').exists():
        templates.extend([
            {
                'name': 'Шаблон для вывода всех поставок за последнюю неделю',
                'query': 'SELECT * FROM core_delivery WHERE date >= date(\'now\', \'-7 days\')'
            },
            {
                'name': 'Шаблон для вывода продуктов с низким остатком',
                'query': 'SELECT * FROM core_product WHERE remaining_stock < 10 ORDER BY remaining_stock ASC'
            },
            {
                'name': 'Шаблон для вывода всех заявок за последний месяц',
                'query': 'SELECT * FROM core_request WHERE date >= date(\'now\', \'-30 days\')'
            }
        ])
    elif user.groups.filter(name='Менеджер по кадрам').exists():
        templates.extend([
            {
                'name': 'Шаблон для вывода всех сотрудников определенной должности',
                'query': 'SELECT * FROM core_employee WHERE position_id = (SELECT id FROM core_position WHERE name = \'Повар\')'
            },
            {
                'name': 'Шаблон для вывода сотрудников по городу',
                'query': 'SELECT e.first_name, e.last_name, c.name as city FROM core_employee e JOIN core_city c ON e.city_id = c.id WHERE c.name = \'Москва\''
            },
            {
                'name': 'Шаблон для вывода сотрудников с определенной профессией',
                'query': 'SELECT * FROM core_employee e JOIN core_workbook w ON e.id = w.employee_id JOIN core_profession p ON w.profession_id = p.id WHERE p.name LIKE \'%повар%\''
            }
        ])
    elif user.groups.filter(name='Директор').exists() or user.is_superuser:
        templates.extend([
            {
                'name': 'Шаблон для вывода всех блюд с ценой выше средней',
                'query': 'SELECT * FROM core_dish WHERE price > (SELECT AVG(price) FROM core_dish)'
            },
            {
                'name': 'Шаблон для вывода топ-5 самых дорогих блюд',
                'query': 'SELECT * FROM core_dish ORDER BY price DESC LIMIT 5'
            },
            {
                'name': 'Шаблон для вывода сотрудников по должности',
                'query': 'SELECT e.first_name, e.last_name, p.name as position FROM core_employee e JOIN core_position p ON e.position_id = p.id WHERE p.name LIKE \'%повар%\''
            },
            {
                'name': 'Шаблон для вывода всех поставок по определенному поставщику',
                'query': 'SELECT d.date, pr.name as provider, dp.quantity FROM core_delivery d JOIN core_provider pr ON d.provider_id = pr.id JOIN core_deliveryproduct dp ON d.id = dp.delivery_id WHERE pr.name LIKE \'%поставщик%\''
            },
            {
                'name': 'Шаблон для вывода отчетов за определенный период',
                'query': 'SELECT * FROM core_report WHERE date BETWEEN \'2025-01-01\' AND \'2025-12-31\''
            }
        ])
    
    return templates

def is_valid_select_query(sql_query):
    """Проверяет, что запрос является SELECT запросом"""
    sql_query = sql_query.strip().upper()
    return sql_query.startswith('SELECT ')

def is_query_using_allowed_tables(sql_query, allowed_models):
    """Проверяет, что запрос использует только разрешенные таблицы"""
    
    allowed_table_names = [model['name'] for model in allowed_models]

    intermediate_tables = []
    for model_dict in allowed_models:
        model_name = model_dict['name']
        # Пример: для модели dish добавляем таблицу dish_ingredients
        if model_name == 'dish':
            intermediate_tables.extend(['dish_ingredients'])
        elif model_name == 'delivery':
            intermediate_tables.extend(['deliveryproduct'])
        elif model_name == 'request':
            intermediate_tables.extend(['requestproduct'])
        elif model_name == 'report':
            intermediate_tables.extend(['reportdish'])
        elif model_name == 'workbook':
            intermediate_tables.extend(['workbook_employees'])
        # Добавьте другие промежуточные таблицы по необходимости
    
    # Объединяем разрешенные таблицы
    all_allowed_tables = allowed_table_names + intermediate_tables
    
    # Паттерн для поиска имен таблиц после FROM, JOIN и т.д.
    pattern = r'(?:FROM|JOIN|INNER\s+JOIN|LEFT\s+JOIN|RIGHT\s+JOIN|CROSS\s+JOIN)\s+([`"]?)([a-zA-Z_][a-zA-Z0-9_]*)(\1)(?:\s|$|[^a-zA-Z0-9_])'
    
    matches = re.findall(pattern, sql_query.upper())
    
    found_tables = []
    for match in matches:
        table_name = match[1]  # Второй элемент - это имя таблицы
        # Убираем префикс 'core_' и 'auth_' и другие префиксы
        clean_table = table_name.replace('CORE_', '').replace('AUTH_', '').replace('SESSIONS_', '').replace('ADMIN_', '').replace('CONTENTTYPES_', '')
        if clean_table:
            found_tables.append(clean_table.lower())
    
    for table in found_tables:
        # Проверяем, есть ли таблица в разрешенных моделях
        if table not in all_allowed_tables:
            # Также проверяем, может быть это системная таблица Django
            if table in ['auth_user', 'auth_group', 'auth_permission', 'django_admin_log', 'django_content_type',
                         'django_session', 'django_migrations', 'auth_user_groups', 'auth_user_user_permissions',
                         'django_admin_log', 'auth_group_permissions', 'django_content_type']:
                return False
            # Проверяем, начинается ли имя с системных префиксов
            if any(table.startswith(prefix) for prefix in ['auth_', 'session_', 'admin_', 'contenttype_', 'logentry_', 'permission_', 'group_', 'user_']):
                return False
        # Проверяем, что таблица в списке разрешенных
        if table not in all_allowed_tables:
            return False
    
    return True